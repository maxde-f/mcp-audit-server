"""
Export Service Executor
Exports reports to PDF, Excel, and other formats
"""
import json
import os
from datetime import datetime
from typing import Any, Dict
from pathlib import Path
import logging

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from base import BaseExecutor

logger = logging.getLogger(__name__)


class ExportServiceExecutor(BaseExecutor):
    """Executor for exporting reports to various formats"""

    name = "export-service"

    def __init__(self, config: Dict[str, Any] = None):
        super().__init__(config)
        self.output_dir = Path(config.get("output_dir", "/tmp/audit-exports") if config else "/tmp/audit-exports")
        self.output_dir.mkdir(parents=True, exist_ok=True)

    async def run(self, action: str, inputs: Dict[str, Any]) -> Dict[str, Any]:
        if action == "export":
            return await self.export(**inputs)
        raise ValueError(f"Unknown action: {action}")

    async def export(
        self,
        report: dict,
        format: str,
        scores: dict,
        cost: dict,
        **kwargs
    ) -> Dict[str, Any]:
        """
        Export report to specified format.

        Args:
            report: Report data from report-generator
            format: Export format (json, pdf, excel, markdown)
            scores: Scoring results
            cost: Cost estimation
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        analysis_id = report.get("analysis_id", "unknown")

        if format == "json":
            return await self._export_json(report, scores, cost, timestamp, analysis_id)
        elif format == "pdf":
            return await self._export_pdf(report, scores, cost, timestamp, analysis_id)
        elif format == "excel":
            return await self._export_excel(report, scores, cost, timestamp, analysis_id)
        elif format == "markdown":
            return await self._export_markdown(report, scores, cost, timestamp, analysis_id)
        else:
            raise ValueError(f"Unsupported format: {format}")

    async def _export_json(
        self,
        report: dict,
        scores: dict,
        cost: dict,
        timestamp: str,
        analysis_id: str
    ) -> Dict[str, Any]:
        """Export to JSON file"""
        filename = f"audit_{analysis_id}_{timestamp}.json"
        filepath = self.output_dir / filename

        export_data = {
            "analysis_id": analysis_id,
            "exported_at": datetime.now().isoformat(),
            "scores": scores,
            "cost": cost,
            "report": report
        }

        with open(filepath, "w") as f:
            json.dump(export_data, f, indent=2, default=str)

        return {
            "file_path": str(filepath),
            "file_url": f"file://{filepath}",
            "file_format": "json",
            "file_size": filepath.stat().st_size
        }

    async def _export_pdf(
        self,
        report: dict,
        scores: dict,
        cost: dict,
        timestamp: str,
        analysis_id: str
    ) -> Dict[str, Any]:
        """Export to PDF file"""
        filename = f"audit_{analysis_id}_{timestamp}.pdf"
        filepath = self.output_dir / filename

        # Try to use available PDF library
        try:
            pdf_content = self._generate_pdf_content(report, scores, cost)
            await self._write_pdf(filepath, pdf_content)
        except ImportError:
            # Fallback: generate HTML that can be printed to PDF
            html_content = self._generate_html_report(report, scores, cost)
            html_path = self.output_dir / f"audit_{analysis_id}_{timestamp}.html"
            with open(html_path, "w") as f:
                f.write(html_content)

            return {
                "file_path": str(html_path),
                "file_url": f"file://{html_path}",
                "file_format": "html",
                "message": "PDF library not available, exported as HTML",
                "file_size": html_path.stat().st_size
            }

        return {
            "file_path": str(filepath),
            "file_url": f"file://{filepath}",
            "file_format": "pdf",
            "file_size": filepath.stat().st_size if filepath.exists() else 0
        }

    async def _export_excel(
        self,
        report: dict,
        scores: dict,
        cost: dict,
        timestamp: str,
        analysis_id: str
    ) -> Dict[str, Any]:
        """Export to Excel file"""
        filename = f"audit_{analysis_id}_{timestamp}.xlsx"
        filepath = self.output_dir / filename

        try:
            import openpyxl
            wb = openpyxl.Workbook()

            # Summary sheet
            ws = wb.active
            ws.title = "Summary"
            self._write_excel_summary(ws, scores, cost)

            # Scores sheet
            ws_scores = wb.create_sheet("Scores")
            self._write_excel_scores(ws_scores, scores)

            # Cost sheet
            ws_cost = wb.create_sheet("Cost Estimation")
            self._write_excel_cost(ws_cost, cost)

            wb.save(filepath)

        except ImportError:
            # Fallback: export as CSV
            csv_path = self.output_dir / f"audit_{analysis_id}_{timestamp}.csv"
            self._write_csv_report(csv_path, scores, cost)

            return {
                "file_path": str(csv_path),
                "file_url": f"file://{csv_path}",
                "file_format": "csv",
                "message": "openpyxl not available, exported as CSV",
                "file_size": csv_path.stat().st_size
            }

        return {
            "file_path": str(filepath),
            "file_url": f"file://{filepath}",
            "file_format": "xlsx",
            "file_size": filepath.stat().st_size
        }

    async def _export_markdown(
        self,
        report: dict,
        scores: dict,
        cost: dict,
        timestamp: str,
        analysis_id: str
    ) -> Dict[str, Any]:
        """Export to Markdown file"""
        filename = f"audit_{analysis_id}_{timestamp}.md"
        filepath = self.output_dir / filename

        md_content = report.get("summary_md", self._generate_markdown(scores, cost))

        with open(filepath, "w") as f:
            f.write(md_content)

        return {
            "file_path": str(filepath),
            "file_url": f"file://{filepath}",
            "file_format": "markdown",
            "file_size": filepath.stat().st_size
        }

    def _generate_pdf_content(self, report: dict, scores: dict, cost: dict) -> bytes:
        """Generate PDF content"""
        # Would use reportlab or weasyprint in production
        raise ImportError("PDF generation requires reportlab or weasyprint")

    async def _write_pdf(self, filepath: Path, content: bytes):
        """Write PDF to file"""
        with open(filepath, "wb") as f:
            f.write(content)

    def _generate_html_report(self, report: dict, scores: dict, cost: dict) -> str:
        """Generate HTML report"""
        readiness = scores.get("overall_readiness", 0)
        level = scores.get("product_level", "Unknown")

        return f"""<!DOCTYPE html>
<html>
<head>
    <title>Audit Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; }}
        h1 {{ color: #333; }}
        .score {{ font-size: 24px; font-weight: bold; }}
        .good {{ color: #22c55e; }}
        .warning {{ color: #f59e0b; }}
        .danger {{ color: #ef4444; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
        th {{ background: #f5f5f5; }}
    </style>
</head>
<body>
    <h1>Repository Audit Report</h1>

    <h2>Overall Readiness: <span class="score {'good' if readiness >= 70 else 'warning' if readiness >= 40 else 'danger'}">{readiness}%</span></h2>

    <p>Product Level: <strong>{level}</strong></p>

    <h3>Scores</h3>
    <table>
        <tr><th>Metric</th><th>Score</th><th>Max</th></tr>
        <tr><td>Repository Health</td><td>{scores.get('repo_health', 0)}</td><td>12</td></tr>
        <tr><td>Technical Debt</td><td>{scores.get('tech_debt', 0)}</td><td>15</td></tr>
        <tr><td>Security</td><td>{scores.get('security_score', 0)}</td><td>3</td></tr>
    </table>

    <h3>Cost Estimation</h3>
    <table>
        <tr><th>Metric</th><th>Value</th></tr>
        <tr><td>Estimated Hours</td><td>{cost.get('estimated_hours', 'N/A')}</td></tr>
        <tr><td>Estimated Cost (USD)</td><td>${cost.get('estimated_cost_usd', 0):,.0f}</td></tr>
        <tr><td>Timeline (weeks)</td><td>{cost.get('timeline_weeks', 'N/A')}</td></tr>
        <tr><td>Team Size</td><td>{cost.get('team_size_recommended', 'N/A')}</td></tr>
    </table>

    <p><em>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}</em></p>
</body>
</html>"""

    def _write_excel_summary(self, ws, scores: dict, cost: dict):
        """Write summary sheet"""
        ws['A1'] = "Metric"
        ws['B1'] = "Value"

        data = [
            ("Overall Readiness", f"{scores.get('overall_readiness', 0)}%"),
            ("Product Level", scores.get('product_level', 'Unknown')),
            ("Repository Health", f"{scores.get('repo_health', 0)}/12"),
            ("Technical Debt", f"{scores.get('tech_debt', 0)}/15"),
            ("Security Score", f"{scores.get('security_score', 0)}/3"),
            ("", ""),
            ("Estimated Hours", cost.get('estimated_hours', 'N/A')),
            ("Estimated Cost (USD)", f"${cost.get('estimated_cost_usd', 0):,.0f}"),
            ("Timeline (weeks)", cost.get('timeline_weeks', 'N/A')),
        ]

        for i, (metric, value) in enumerate(data, start=2):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value

    def _write_excel_scores(self, ws, scores: dict):
        """Write scores sheet"""
        ws['A1'] = "Metric"
        ws['B1'] = "Score"
        ws['C1'] = "Max"

        breakdown = scores.get('breakdown', {})
        health = breakdown.get('health', {})
        debt = breakdown.get('debt', {})

        row = 2
        for metric, data in health.items():
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = data.get('points', data.get('present', 'N/A'))
            ws[f'C{row}'] = data.get('max', '')
            row += 1

    def _write_excel_cost(self, ws, cost: dict):
        """Write cost sheet"""
        ws['A1'] = "Metric"
        ws['B1'] = "Value"

        row = 2
        for key, value in cost.items():
            if not isinstance(value, dict):
                ws[f'A{row}'] = key
                ws[f'B{row}'] = str(value)
                row += 1

    def _write_csv_report(self, filepath: Path, scores: dict, cost: dict):
        """Write CSV fallback"""
        with open(filepath, 'w') as f:
            f.write("Metric,Value\n")
            f.write(f"Overall Readiness,{scores.get('overall_readiness', 0)}%\n")
            f.write(f"Product Level,{scores.get('product_level', 'Unknown')}\n")
            f.write(f"Repository Health,{scores.get('repo_health', 0)}/12\n")
            f.write(f"Technical Debt,{scores.get('tech_debt', 0)}/15\n")
            f.write(f"Security Score,{scores.get('security_score', 0)}/3\n")
            f.write(f"Estimated Hours,{cost.get('estimated_hours', 'N/A')}\n")
            f.write(f"Estimated Cost USD,{cost.get('estimated_cost_usd', 0)}\n")

    def _generate_markdown(self, scores: dict, cost: dict) -> str:
        """Generate markdown report"""
        return f"""# Audit Report

## Summary

- **Overall Readiness**: {scores.get('overall_readiness', 0)}%
- **Product Level**: {scores.get('product_level', 'Unknown')}

## Scores

| Metric | Score | Max |
|--------|-------|-----|
| Repository Health | {scores.get('repo_health', 0)} | 12 |
| Technical Debt | {scores.get('tech_debt', 0)} | 15 |
| Security | {scores.get('security_score', 0)} | 3 |

## Cost Estimation

- **Hours**: {cost.get('estimated_hours', 'N/A')}
- **Cost (USD)**: ${cost.get('estimated_cost_usd', 0):,.0f}
- **Timeline**: {cost.get('timeline_weeks', 'N/A')} weeks
- **Team Size**: {cost.get('team_size_recommended', 'N/A')}

---
*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*
"""

    def get_capabilities(self) -> list[str]:
        return ["export"]


def create_executor(config: Dict[str, Any] = None) -> ExportServiceExecutor:
    return ExportServiceExecutor(config)
